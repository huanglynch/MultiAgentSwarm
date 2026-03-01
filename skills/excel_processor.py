# skills/excel_processor.py
# 全能 Excel 处理 Skill - 完全兼容当前 MultiAgentSwarm 架构
# 参考 PDF_Read 格式：单个 tool_function + tool_schema
# 支持：读取、分析、编辑、写入、创建新文件，返回 Markdown 表格（WebUI 美观显示）

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook


def tool_function(
        file_path: str,
        operation: str = "read",  # read / analyze / edit / write / create
        sheet_name: str = None,
        cell: str = None,  # 编辑用：如 "A1"
        value: any = None,  # 编辑用：新值
        data: dict = None,  # write/create 用（可直接传列表或 dict）
        output_filename: str = None  # 新文件名（可选）
) -> dict:
    """
    全能 Excel 处理器（仅限 uploads/ 目录）
    """
    try:
        path = Path(file_path)
        upload_dir = Path("uploads")

        # 安全铁律：强制 uploads/ 目录
        if not str(path.resolve()).startswith(str(upload_dir.resolve())):
            return {"success": False, "error": "安全铁律：仅允许处理 uploads/ 目录下的文件"}

        result = {"success": True, "operation": operation, "file_path": str(path)}

        if operation in ["read", "analyze"]:
            if not path.exists():
                return {"success": False, "error": f"文件不存在: {file_path}"}

            excel_file = pd.ExcelFile(path)
            result["sheets"] = excel_file.sheet_names

            target_sheet = sheet_name or excel_file.sheet_names[0]
            df = pd.read_excel(path, sheet_name=target_sheet)

            # 结构化数据（前 200 行防超长）
            result["data"] = df.head(200).to_dict(orient="records")
            result["shape"] = [len(df), len(df.columns)]
            result["columns"] = list(df.columns)

            # Markdown 表格（WebUI 直接美观显示）
            result["markdown"] = df.head(30).to_markdown(index=False)

            if operation == "analyze":
                result["analysis"] = {
                    "row_count": len(df),
                    "col_count": len(df.columns),
                    "numeric_cols": list(df.select_dtypes(include="number").columns),
                    "sample": df.head(5).to_dict(orient="records")
                }

        elif operation == "edit":
            if not path.exists() or not cell or value is None:
                return {"success": False, "error": "edit 操作需要有效 cell 和 value"}
            wb = load_workbook(path)
            ws = wb[sheet_name or wb.sheetnames[0]]
            ws[cell] = value
            wb.save(path)
            result["message"] = f"单元格 {cell} 已更新为 {value}"

        elif operation in ["write", "create"]:
            if not data:
                return {"success": False, "error": "write/create 需要 data 参数"}

            # 支持 list 或 dict 两种格式
            if isinstance(data, list):
                df = pd.DataFrame(data)
            else:
                df = pd.DataFrame.from_dict(data, orient="index") if isinstance(data, dict) else pd.DataFrame(data)

            output_path = upload_dir / (output_filename or f"new_{path.name}" if operation == "create" else path.name)
            df.to_excel(output_path, index=False)
            result["saved_path"] = str(output_path)
            result["markdown"] = df.head(20).to_markdown(index=False)
            result["message"] = f"文件已保存：{output_path.name}"

        return result

    except Exception as e:
        return {"success": False, "error": str(e)}


# ====================== Tool Schema ======================
tool_schema = {
    "type": "function",
    "function": {
        "name": "excel_processor",
        "description": "全能 Excel 文档处理/编辑/读写 Skill（仅限 uploads/）。支持 read/analyze/edit/write/create，返回 Markdown 表格 + 结构化数据。推荐优先使用 analyze 或 read 操作。",
        "parameters": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Excel 文件路径（必须在 uploads/ 下）"},
                "operation": {"type": "string", "enum": ["read", "analyze", "edit", "write", "create"],
                              "default": "read"},
                "sheet_name": {"type": "string", "description": "指定工作表（可选，默认第一个）"},
                "cell": {"type": "string", "description": "edit 操作：单元格坐标，如 A1"},
                "value": {"type": "string", "description": "edit 操作：新值"},
                "data": {"type": "object", "description": "write/create 操作：数据（列表或 dict）"},
                "output_filename": {"type": "string", "description": "新文件名（可选）"}
            },
            "required": ["file_path"]
        }
    }
}